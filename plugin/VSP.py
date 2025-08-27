import openpyxl


outfile = open('./VSP_accession.list', "w")
def parse_microorganisms_by_column_name(file_path):
    """
    解析VSPv2_P-R_Panel_summary.xlsx文件，根据列名提取"Microorganisms"工作表中的数据。

    Args:
        file_path (str): XLSX文件的完整路径。

    Returns:
        dict: 键为Reporting Name，值为一个包含Genome Length Reference元素的列表。
              如果文件或工作表不存在，或找不到指定列，返回None。
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
    except FileNotFoundError:
        print(f"错误：文件 '{file_path}' 不存在。")
        return None

    if 'Microorganisms' not in workbook.sheetnames:
        print(f"错误：文件 '{file_path}' 中不存在名为 'Microorganisms' 的工作表。")
        return None

    sheet = workbook['Microorganisms']

    # 1. 查找表头，确定列索引
    header = [cell.value for cell in sheet[1]]  # 假设第一行是表头
    # 查找所需列的索引
    try:
        reporting_name_col_index = header.index('Reporting Name')
        genome_length_ref_col_index = header.index('Genome Length Reference')
    except ValueError as e:
        print(f"错误：找不到指定的列名。详细信息：{e}")
        return None

    microorganism_data = {}
    accession=[]
    # 2. 遍历数据行，使用确定的索引提取数据
    for row in sheet.iter_rows(min_row=2):  # 从第二行开始遍历数据
        try:
            # 使用索引获取单元格的值
            reporting_name = row[reporting_name_col_index].value
            genome_length_reference = row[genome_length_ref_col_index].value

            # 检查并处理数据
            if reporting_name and genome_length_reference and str(genome_length_reference).strip() != '-':
                reference_list = [item.strip() for item in str(genome_length_reference).split(',')]
                for item in str(genome_length_reference).split(','):
                    outfile.write(f'{item}\n')
                microorganism_data[reporting_name] = reference_list

        except IndexError:
            # 捕获因某些行单元格数量不足而导致的索引错误
            continue

    workbook.close()
    outfile.close()


# 使用示例
if __name__ == "__main__":
    excel_file_path = './VSPV2_2-7-0_Panel_Summary.xlsx'
    parsed_data = parse_microorganisms_by_column_name(excel_file_path)